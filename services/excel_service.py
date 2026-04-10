"""
Nexus AI — Excel Service
Генерация .xlsx файлов с данными о мероприятии
"""
import os
import re
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ─── Helpers ────────────────────────────────────────────────────────────────

def sanitize_for_excel(value):
    """Удаляет управляющие символы, которые ломают openpyxl."""
    if not isinstance(value, str):
        return value
    return re.sub(r'[\000-\010\013\014\016-\037]', '', str(value))


def extract_payment_value(payment_str) -> float:
    """
    Извлекает числовое значение из строки оплаты.
    Примеры:
        "4000 ₽" → 4000
        "350-400/час" → 375 (среднее)
        "5000" → 5000
        "По договоренности" → 0
    """
    if not payment_str:
        return 0.0

    text = str(payment_str).strip()

    # Попытка найти диапазон (350-400)
    range_match = re.search(r'(\d+)\s*[-–]\s*(\d+)', text)
    if range_match:
        low = float(range_match.group(1))
        high = float(range_match.group(2))
        return (low + high) / 2

    # Попытка найти одно число
    num_match = re.search(r'(\d+(?:[.,]\d+)?)', text)
    if num_match:
        return float(num_match.group(1).replace(',', '.'))

    return 0.0


def calc_hours(arrival, departure) -> float:
    """Расчёт отработанных часов между arrival и departure (HH:MM)."""
    if not arrival or not departure:
        return 0.0

    arrival_s = str(arrival).strip().replace('\u2014', '-').replace('\u2013', '-')
    departure_s = str(departure).strip().replace('\u2014', '-').replace('\u2013', '-')

    if arrival_s in ('—', '–', '—', 'не указано', '') or \
       departure_s in ('—', '–', '—', 'не указано', ''):
        return 0.0

    try:
        fmt = "%H:%M"
        t1 = datetime.strptime(arrival_s, fmt)
        t2 = datetime.strptime(departure_s, fmt)
        delta = t2 - t1
        if delta.total_seconds() < 0:
            return round((delta.total_seconds() + 86400) / 3600, 2)
        return round(delta.total_seconds() / 3600, 2)
    except Exception as e:
        logger.debug(f"Error calculating hours for {arrival_s}-{departure_s}: {e}")
        return 0.0


def status_label(status: str) -> str:
    """Человекочитаемый статус заявки."""
    mapping = {
        "PENDING": "Ожидает",
        "ACCEPTED": "Принят",
        "SCHEDULED": "Назначен",
        "INVITED": "Приглашён",
        "CONFIRMED": "Подтверждён",
        "CHECKED_IN": "Отметился",
        "REJECTED": "Отклонён",
        "DECLINED": "Отказался",
    }
    return mapping.get(status, status or "—")


# Стили
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
SUMMARY_FONT = Font(bold=True, size=12)
SUMMARY_FILL = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin'),
)

def style_header_row(ws, row_num: int, num_cols: int):
    """Применяет стили к строке заголовков."""
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_range(ws, start_row: int, end_row: int, num_cols: int):
    """Применяет границы ко всем ячейкам данных."""
    for r in range(start_row, end_row + 1):
        for c in range(1, num_cols + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER


def auto_width(ws, max_row: int, max_col: int):
    """Авто-ширина колонок."""
    for col_idx in range(1, max_col + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row_idx in range(1, max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        ws.column_dimensions[letter].width = min(max_len + 4, 40)


# ─── Основной отчёт по мероприятию ──────────────────────────────────────────

def generate_event_xlsx(
    event_title: str,
    event_date: str,
    event_location: str,
    candidates: list,
    event_payment: str = None,
    event_status: str = None,
    event_created_at: str = None,
    required_men: int = 0,
    required_women: int = 0,
) -> str:
    """
    Генерирует детальный отчёт по одному мероприятию.
    Включает листы: Кандидаты, Сводка.
    """
    wb = Workbook()
    rate = extract_payment_value(event_payment)

    # ════════ Лист 1: Кандидаты ════════
    ws = wb.active
    ws.title = "Кандидаты"

    # Заголовок мероприятия
    ws.merge_cells("A1:J1")
    ws["A1"] = f"Мероприятие: {event_title} ({event_date}) — {event_location}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Шапка: ФИО | Пол | Телефон | TG | Роль | Приход | Уход | Часы | Статус | Ставка | К оплате
    headers = ["ФИО", "Пол", "Телефон", "TG Username", "Роль", "Приход", "Уход",
               "Отработано (ч)", "Статус", "Ставка", "К оплате"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    total_payable = 0.0
    row_idx = 3
    for c in candidates:
        profile = c.get("candidates", {}) or {}
        full_name = sanitize_for_excel(
            profile.get("full_name") or
            f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip() or
            "—"
        )
        gender_raw = profile.get("gender", "")
        gender = "Мужской" if gender_raw == "Male" else "Женский" if gender_raw == "Female" else "—"
        phone = sanitize_for_excel(profile.get("phone_number") or "—")
        username = sanitize_for_excel(profile.get("telegram_username") or "—")
        role = sanitize_for_excel(c.get("role") or profile.get("primary_role") or "—")
        arrival = c.get("arrival_time") or "—"
        departure = c.get("departure_time") or "—"
        app_status = c.get("application_status", "")
        status_text = status_label(app_status)

        hours = calc_hours(arrival, departure)
        payable = round(hours * rate, 2) if hours > 0 and rate > 0 else 0
        total_payable += payable

        ws.append([
            full_name, gender, phone, username, role,
            arrival, departure,
            hours if hours > 0 else "—",
            status_text,
            rate if rate > 0 else "—",
            payable if payable > 0 else "—"
        ])

        row_idx += 1

    # Итоговая строка
    ws.append(["", "", "", "", "", "", "", "", f"ИТОГО ({row_idx - 3} чел.):", "", total_payable if total_payable > 0 else "—"])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    style_data_range(ws, 2, row_idx, len(headers))
    auto_width(ws, row_idx, len(headers))

    # ════════ Лист 2: Сводка ════════
    ws2 = wb.create_sheet("Сводка")

    total_applicants = len(candidates)
    accepted = sum(1 for c in candidates if c.get("application_status") == "ACCEPTED")
    rejected = sum(1 for c in candidates if c.get("application_status") == "REJECTED")
    confirmed = sum(1 for c in candidates if c.get("application_status") == "CONFIRMED")
    checked_in = sum(1 for c in candidates if c.get("application_status") == "CHECKED_IN")
    men_count = sum(1 for c in candidates if (c.get("candidates") or {}).get("gender") == "Male")
    women_count = sum(1 for c in candidates if (c.get("candidates") or {}).get("gender") == "Female")

    summary_data = [
        ("Мероприятие", event_title),
        ("Дата", event_date),
        ("Место", event_location),
        ("Оплата", event_payment or "—"),
        ("Ставка/час", f"{rate} ₽" if rate > 0 else "—"),
        ("", ""),
        ("Всего заявок", total_applicants),
        ("Принято", accepted),
        ("Отклонено", rejected),
        ("Подтверждено", confirmed),
        ("Отметились", checked_in),
        ("", ""),
        ("Нужно мужчин", required_men),
        ("Нужно женщин", required_women),
        ("", ""),
        ("Мужчин откликнулось", men_count),
        ("Женщин откликнулось", women_count),
        ("", ""),
        ("Общий фонд оплаты", f"{total_payable:,.0f} ₽" if total_payable > 0 else "—"),
    ]

    if event_created_at:
        # Расчёт Time to Hire: от создания мероприятия до первого отклика
        try:
            created_dt = datetime.fromisoformat(event_created_at.replace("Z", "+00:00").replace("+00:00", ""))
            # Ищем первый лог "Candidate Registered"
            # Эта информация передаётся через event_logs отдельно
            summary_data.insert(4, ("Создано", event_created_at[:19]))
        except Exception:
            pass

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "Сводка по мероприятию"
    ws2["A1"].font = Font(bold=True, size=14)

    row = 3
    for label, value in summary_data:
        ws2.cell(row=row, column=1, value=label).font = Font(bold=True, size=11)
        ws2.cell(row=row, column=2, value=value).font = Font(size=11)
        row += 1

    ws2.column_dimensions["A"].width = 30
    ws2.column_dimensions["B"].width = 35

    # Сохранение
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"event_{event_title[:20]}_{timestamp}.xlsx".replace(" ", "_")
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    logger.info(f"Excel report saved: {filepath}")
    return filepath


# ─── Общий отчёт по компании ────────────────────────────────────────────────

def generate_company_report_xlsx(company_name: str, candidates: list) -> str:
    """
    Генерирует общий отчёт по всем мероприятиям компании.
    Включает листы: Все заявки, Помесячная сводка.
    """
    wb = Workbook()

    # ════════ Лист 1: Все заявки ════════
    ws = wb.active
    ws.title = "Все заявки"

    ws.merge_cells("A1:L1")
    ws["A1"] = f"Общий отчёт по компании: {company_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Проект", "Дата", "ФИО", "Пол", "Телефон", "TG Username",
               "Роль", "Приход", "Уход", "Отработано (ч)", "Статус", "Оплата"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    total_payable = 0.0
    row_idx = 3

    for c in candidates:
        profile = c.get("candidates", {}) or {}
        event = c.get("events", {}) or {}

        full_name = profile.get("full_name") or f"{profile.get('first_name', '')} {profile.get('last_name', '')}".strip()
        gender_raw = profile.get("gender", "")
        gender = "Мужской" if gender_raw == "Male" else "Женский" if gender_raw == "Female" else "—"

        arrival = c.get("arrival_time") or "—"
        departure = c.get("departure_time") or "—"
        hours = calc_hours(arrival, departure)

        # Извлекаем ставку из оплаты мероприятия
        event_payment = event.get("payment", "")
        rate = extract_payment_value(event_payment)
        payable = round(hours * rate, 2) if hours > 0 and rate > 0 else 0
        total_payable += payable

        app_status = c.get("application_status", "")
        role = c.get("role", profile.get("primary_role", "—"))

        ws.append([
            sanitize_for_excel(event.get("title", "—")),
            sanitize_for_excel(str(event.get("date", "—"))[:10]),
            sanitize_for_excel(full_name or "—"),
            gender,
            sanitize_for_excel(profile.get("phone_number") or "—"),
            sanitize_for_excel(profile.get("telegram_username") or "—"),
            sanitize_for_excel(role),
            arrival,
            departure,
            hours if hours > 0 else "—",
            status_label(app_status),
            payable if payable > 0 else "—",
        ])
        row_idx += 1

    # Итоговая строка
    ws.append(["", "", "", "", "", "", "", "", "", "", f"ИТОГО ({row_idx - 3} чел.):", total_payable if total_payable > 0 else "—"])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    style_data_range(ws, 2, row_idx, len(headers))
    auto_width(ws, row_idx, len(headers))

    # ════════ Лист 2: Помесячная сводка ════════
    ws2 = wb.create_sheet("Помесячная сводка")

    # Агрегация по месяцам
    monthly: dict[str, dict] = {}
    for c in candidates:
        event = c.get("events", {}) or {}
        date_str = str(event.get("date", ""))[:10]
        if not date_str or date_str == "—":
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            month_key = dt.strftime("%Y-%m")
            month_label = dt.strftime("%B %Y")
        except ValueError:
            continue

        if month_key not in monthly:
            monthly[month_key] = {
                "label": month_label,
                "events": set(),
                "applicants": 0,
                "accepted": 0,
                "total_hours": 0.0,
                "total_payable": 0.0,
            }

        ev_id = event.get("event_id", "")
        monthly[month_key]["events"].add(ev_id)
        monthly[month_key]["applicants"] += 1
        if c.get("application_status") == "ACCEPTED":
            monthly[month_key]["accepted"] += 1

        arrival = c.get("arrival_time") or "—"
        departure = c.get("departure_time") or "—"
        hours = calc_hours(arrival, departure)
        event_payment = event.get("payment", "")
        rate = extract_payment_value(event_payment)
        pay = round(hours * rate, 2) if hours > 0 and rate > 0 else 0

        monthly[month_key]["total_hours"] += hours
        monthly[month_key]["total_payable"] += pay

    # Заголовки
    monthly_headers = ["Месяц", "Мероприятий", "Всего откликов", "Принято", "Отработано (ч)", "Фонд оплаты"]
    ws2.append(monthly_headers)
    style_header_row(ws2, 1, len(monthly_headers))

    row = 2
    grand_events = 0
    grand_applicants = 0
    grand_accepted = 0
    grand_hours = 0.0
    grand_payable = 0.0

    for month_key in sorted(monthly.keys()):
        m = monthly[month_key]
        num_events = len(m["events"])
        grand_events += num_events
        grand_applicants += m["applicants"]
        grand_accepted += m["accepted"]
        grand_hours += m["total_hours"]
        grand_payable += m["total_payable"]

        ws2.append([
            m["label"],
            num_events,
            m["applicants"],
            m["accepted"],
            round(m["total_hours"], 1),
            round(m["total_payable"], 2),
        ])
        for col in range(1, len(monthly_headers) + 1):
            ws2.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    # Итого
    ws2.append([
        "ИТОГО",
        grand_events,
        grand_applicants,
        grand_accepted,
        round(grand_hours, 1),
        round(grand_payable, 2),
    ])
    for col in range(1, len(monthly_headers) + 1):
        cell = ws2.cell(row=row, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 18
    ws2.column_dimensions["D"].width = 12
    ws2.column_dimensions["E"].width = 18
    ws2.column_dimensions["F"].width = 18

    # ════════ Лист 3: Метрика Time to Hire ════════
    ws3 = wb.create_sheet("Time to Hire")

    ws3.merge_cells("A1:C1")
    ws3["A1"] = "Метрика Time to Hire — от публикации до набора"
    ws3["A1"].font = Font(bold=True, size=14)

    tth_headers = ["Мероприятие", "Дата мероприятия", "Опубликовано", "Первый отклик", "Time to Hire (дней)"]
    ws3.append(tth_headers)
    style_header_row(ws3, 2, len(tth_headers))

    # Агрегация по мероприятиям
    events_data: dict[str, dict] = {}
    for c in candidates:
        event = c.get("events", {}) or {}
        ev_id = event.get("event_id", "")
        if not ev_id:
            continue

        if ev_id not in events_data:
            events_data[ev_id] = {
                "title": event.get("title", "—"),
                "date": str(event.get("date", "—"))[:10],
                "created_at": event.get("created_at", ""),
                "first_app": None,
            }

        # Время создания заявки — используем event_logs если доступны,
        # иначе fallback на created_at мероприятия
        log_ts = c.get("_first_log_timestamp")
        if log_ts:
            if events_data[ev_id]["first_app"] is None or log_ts < events_data[ev_id]["first_app"]:
                events_data[ev_id]["first_app"] = log_ts

    row = 3
    for ev_id, ed in events_data.items():
        # Если нет точного времени первого отклика, используем created_at как приблизительный
        if ed["first_app"]:
            first_app_display = ed["first_app"][:19]
            try:
                created_dt = datetime.fromisoformat(ed["created_at"].replace("Z", "")) if ed["created_at"] else None
                first_dt = datetime.fromisoformat(ed["first_app"].replace("Z", ""))
                if created_dt:
                    days = max(0, (first_dt - created_dt).total_seconds() / 86400)
                    days_display = round(days, 1)
                else:
                    days_display = "—"
            except Exception:
                days_display = "—"
        else:
            first_app_display = "—"
            days_display = "—"

        ws3.append([
            sanitize_for_excel(ed["title"]),
            ed["date"],
            ed["created_at"][:19] if ed["created_at"] else "—",
            first_app_display,
            days_display,
        ])
        for col in range(1, len(tth_headers) + 1):
            ws3.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    ws3.column_dimensions["A"].width = 35
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 22
    ws3.column_dimensions["D"].width = 22
    ws3.column_dimensions["E"].width = 22

    # Сохранение
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"company_{company_name[:20]}_{timestamp}.xlsx".replace(" ", "_")
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    logger.info(f"Company report saved: {filepath}")
    return filepath


# ─── Помесячный отчёт (отдельный файл) ──────────────────────────────────────

def generate_monthly_report_xlsx(company_name: str, candidates: list, events: list = None) -> str:
    """
    Детальный помесячный отчёт: одно мероприятие = одна строка + детали.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Помесячный отчёт"

    ws.merge_cells("A1:G1")
    ws["A1"] = f"Помесячный отчёт: {company_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["Месяц", "Мероприятие", "Дата", "Всего заявок", "Принято", "Отработано (ч)", "Фонд оплаты"]
    ws.append(headers)
    style_header_row(ws, 2, len(headers))

    # Группировка по мероприятиям
    monthly: dict[str, dict] = {}
    for c in candidates:
        event = c.get("events", {}) or {}
        date_str = str(event.get("date", ""))[:10]
        if not date_str or date_str == "—":
            continue
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d")
            month_key = dt.strftime("%Y-%m")
            month_label = dt.strftime("%B %Y")
        except ValueError:
            continue

        ev_id = event.get("event_id", "")

        if ev_id not in monthly:
            monthly[ev_id] = {
                "month_label": month_label,
                "month_key": month_key,
                "title": event.get("title", "—"),
                "date": date_str,
                "applicants": 0,
                "accepted": 0,
                "total_hours": 0.0,
                "total_payable": 0.0,
            }

        monthly[ev_id]["applicants"] += 1
        if c.get("application_status") == "ACCEPTED":
            monthly[ev_id]["accepted"] += 1

        arrival = c.get("arrival_time") or "—"
        departure = c.get("departure_time") or "—"
        hours = calc_hours(arrival, departure)
        event_payment = event.get("payment", "")
        rate = extract_payment_value(event_payment)
        pay = round(hours * rate, 2) if hours > 0 and rate > 0 else 0

        monthly[ev_id]["total_hours"] += hours
        monthly[ev_id]["total_payable"] += pay

    row_idx = 3
    for ev_id in sorted(monthly.keys()):
        m = monthly[ev_id]
        ws.append([
            m["month_label"],
            sanitize_for_excel(m["title"]),
            m["date"],
            m["applicants"],
            m["accepted"],
            round(m["total_hours"], 1),
            round(m["total_payable"], 2),
        ])
        for col in range(1, len(headers) + 1):
            ws3 = ws.cell(row=row_idx, column=col)
            ws3.border = THIN_BORDER
        row_idx += 1

    # Итого
    total_app = sum(m["applicants"] for m in monthly.values())
    total_acc = sum(m["accepted"] for m in monthly.values())
    total_hrs = sum(m["total_hours"] for m in monthly.values())
    total_pay = sum(m["total_payable"] for m in monthly.values())

    ws.append(["", "", "ИТОГО", total_app, total_acc, round(total_hrs, 1), round(total_pay, 2)])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, size=11)
        cell.fill = SUMMARY_FILL
        cell.border = THIN_BORDER

    style_data_range(ws, 2, row_idx, len(headers))
    auto_width(ws, row_idx, len(headers))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"monthly_{company_name[:20]}_{timestamp}.xlsx".replace(" ", "_")
    filepath = os.path.join(os.getcwd(), filename)
    wb.save(filepath)
    return filepath
